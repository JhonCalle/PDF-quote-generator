import os
from openpyxl.worksheet.datavalidation import DataValidation
import getPDFInfo
import openpyxl

def check_num_cotiz(directory):
    # if there are more than 10 excel files, delete the 5 oldest
    excel_files = []
    for file_name in os.listdir(directory):
        if file_name.endswith(".xlsx"):
            excel_files.append(os.path.join(directory, file_name))

    if len(excel_files) > 10:
        # Sort the Excel files by modified time, oldest first
        excel_files.sort(key=lambda x: os.path.getmtime(x))

        # Delete the oldest 5 Excel files
        for i in range(5):
            os.remove(excel_files[i])

    # Find the next consecutive number to use as a filename
    max_number = 0
    for excel_file in excel_files:
        # Extract the number from the filename
        number = int(os.path.basename(excel_file).split('.')[0])
        if number > max_number:
            max_number = number
    return str(max_number + 1)

def inset_data_in_excel(ws, pdf_names, pdf_pages, pdf_size):
    items = len(pdf_names)
    for i in range(items):
        if i != 0 and i!=items-1:
            # If we are not in the first row, we need to insert a row ans copy styles
            ws.insert_rows(7)
            #copy format ans style from row 6
            for j in range(1, 14):
                ws.cell(row=6+i, column=j).value = ws.cell(row=6, column=j).value
                ws.cell(row=6+i, column=j)._style = ws.cell(row=6, column=j)._style

        # Insert data of PDF_info
        ws.cell(row=6+i, column=2).value = i+1
        ws.cell(row=6+i, column=3).value = pdf_names[i]
        ws.cell(row=6+i, column=4).value = pdf_pages[i]
        ws.cell(row=6+i, column=5).value = pdf_size[i]
        ws.cell(row=6+i, column=6).value = "Blanco y negro"
        ws.cell(row=6+i, column=7).value = "=M"+str((6+i))

        # Put formulas to calculate price per page and price per finish
        ws.cell(row=6+i, column=9).value = "Ninguna"




        ws.cell(row=6+i, column=10).value = ("=INDEX(Precios!C5:D7;MATCH(E"+str((6+i))+";Precios!B5:B7;0);MATCH(F"+str((6+i))+";Precios!C4:D4;0))*INDEX(Precios!C10:C12;MATCH(D"+str((6+i))+";Precios!B10:B12;1))*INDEX(Precios!C15:C17;MATCH(I"+str((6+i))+";Precios!B15:B17;0))").replace(";", ",")
        ws.cell(row=6+i, column=11).value = ("=INDEX(Precios!G5:I16;MATCH(D" + str((6 + i)) + ";Precios!F5:F16;1);MATCH(E" + str((6 + i)) + ";Precios!G4:I4;0))").replace(";", ",")

        # Put formulas to calculate the total price
        ws.cell(row=6+i, column=12).value = "=J"+str((6+i))+"*D"+str((6+i))+"+K"+str((6+i))
        ws.cell(row=6+i, column=13).value = "=ROUND(L"+str((6+i))+", 0)"

        #add data validation
        dv_size =DataValidation(type="list", formula1='"Carta, Medio oficio, Oficio"')
        dv_Color = DataValidation(type="list", formula1='"Blanco y negro, Colores"')
        dv_dificulty = DataValidation(type="list", formula1='"Ninguna, Medio, Alta"')
        ws.add_data_validation(dv_size)
        ws.add_data_validation(dv_Color)
        ws.add_data_validation(dv_dificulty)
        dv_size.add(ws.cell(row=6+i, column=5))
        dv_Color.add(ws.cell(row=6+i, column=6))
        dv_dificulty.add(ws.cell(row=6+i, column=9))


    # Make all the final calculations
    ws.cell(row=items+6, column=7).value = "=SUM(G6:G"+str(items+5)+")"
    ws.cell(row=items+8, column=7).value = "=ROUND(G"+str(items+6)+"*G"+str(items+7)+", 0)"
    ws.cell(row=items+9, column=7).value = "=G" + str(items + 6) + "-G" + str(items+8)




#Open a excel file
wb = openpyxl.load_workbook(r"D:\Proyectos Programación\Cotizaciones\Plantilla.xlsx")
ws = wb["Master"]

# insert data in excel
pdf_name, pdf_num_pages, pdf_size = getPDFInfo.get_pdf_info("D:\iB\Trabajos\Por cotizar")
inset_data_in_excel(ws, pdf_name, pdf_num_pages, pdf_size)

#Save the file
name = check_num_cotiz(r"D:\Proyectos Programación\Cotizaciones\Done")
wb.save(r"D:\Proyectos Programación\Cotizaciones\Done\\" + name + ".xlsx")